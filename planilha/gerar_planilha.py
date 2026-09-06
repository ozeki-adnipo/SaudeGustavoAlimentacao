#!/usr/bin/env python3
"""Gera/regenera 'plano-alimentar-pos-infarto.xlsx' a partir deste script.

- Abas 1 e 2 (texto de orientação) ficam definidas diretamente aqui.
- Aba 3 ("Alimentos e Quantidades") é montada a partir de `dados_alimentos.json`,
  que é a fonte de verdade da lista de alimentos.

Para acrescentar um alimento novo: edite `dados_alimentos.json` (adicione um
objeto na lista, respeitando a categoria) e rode este script de novo:

    python3 gerar_planilha.py

Isso recria o arquivo .xlsx inteiro, sempre com a mesma formatação. É assim
que a skill `adicionar-alimento` funciona por baixo dos panos.

Sem fórmulas — é uma planilha de referência/consulta, não um modelo de cálculo.
"""
import json
import math
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName

HERE = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(HERE, "dados_alimentos.json")
OUT_PATH = os.path.join(HERE, "plano-alimentar-pos-infarto.xlsx")

FONT_NAME = "Arial"

# ---- paleta ----
COLOR_TITLE_BG = "1F4E5F"      # azul petróleo escuro (títulos de aba/seção)
COLOR_TITLE_FG = "FFFFFF"
COLOR_SECTION_BG = "D9E6EC"    # azul bem claro (linhas de seção)
COLOR_WARN_BG = "FCE4D6"       # laranja claro (avisos/pendências)
COLOR_LIBERADO = "C6E0B4"      # verde
COLOR_MODERAR = "FFE699"       # amarelo
COLOR_EVITAR = "F4B6B6"        # vermelho claro
COLOR_CATEGORY_BG = "1F4E5F"
COLOR_CATEGORY_FG = "FFFFFF"

CLASS_COLORS = {
    "Liberado": COLOR_LIBERADO,
    "Moderar": COLOR_MODERAR,
    "Evitar": COLOR_EVITAR,
}

# Opções do menu suspenso da coluna "Gostoso" (avaliação pessoal do usuário,
# preenchida à mão na planilha — nunca decidida pela skill adicionar-alimento).
GOSTOSO_OPCOES = ["Ruim", "Normal", "Bom", "Muito bom"]
GOSTOSO_COLORS = {
    "Ruim": COLOR_EVITAR,
    "Normal": COLOR_MODERAR,
    "Bom": COLOR_LIBERADO,
    "Muito bom": "70AD47",  # verde mais forte que "Bom"
}

THIN = Side(style="thin", color="B7B7B7")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ordem fixa das categorias na aba 3. Uma categoria nova (que não esteja
# nesta lista) é acrescentada automaticamente ao final, na ordem em que
# aparecer pela primeira vez no JSON.
CATEGORY_ORDER = [
    "Frutas",
    "Grãos e Cereais",
    "Proteínas",
    "Laticínios",
    "Bebidas",
    "Doces e Sobremesas",
    "Gorduras e Temperos",
    "Vegetais e Verduras",
]

# Refeições usadas na aba "Montar Refeição" — precisam bater exatamente com
# os valores usados no campo "refeicoes" de dados_alimentos.json.
MEALS = ["Café da manhã", "Almoço", "Lanche", "Jantar"]

# Nome definido (Excel "Defined Name") de cada lista de refeição. Necessário
# porque um menu suspenso (validação de dados tipo lista) não pode apontar
# direto para um intervalo em OUTRA aba — só funciona com um nome definido
# ou um intervalo na mesma aba. Sem isso, o menu suspenso simplesmente não
# aparece nem funciona no Excel.
MEAL_DEFINED_NAME = {
    "Café da manhã": "Lista_CafeDaManha",
    "Almoço": "Lista_Almoco",
    "Lanche": "Lista_Lanche",
    "Jantar": "Lista_Jantar",
}

# Para cada categoria, qual categoria costuma combinar bem com ela — usado
# só para calcular a coluna "Sugestão" da aba "Montar Refeição" (uma dica,
# não uma regra rígida). Uma categoria nova criada pela skill
# adicionar-alimento não entra aqui automaticamente — a sugestão para ela
# fica em branco até alguém adicionar uma linha neste dicionário.
CATEGORIA_COMBINA = {
    "Laticínios": "Grãos e Cereais",
    "Grãos e Cereais": "Laticínios",
    "Proteínas": "Vegetais e Verduras",
    "Vegetais e Verduras": "Proteínas",
    "Bebidas": "Proteínas",
    "Frutas": "Laticínios",
    "Doces e Sobremesas": "Frutas",
    "Gorduras e Temperos": "Vegetais e Verduras",
}


def style_title(ws, text, span, row=1, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_TITLE_FG)
    cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height


def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_TITLE_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# Caracteres por linha ≈ largura da coluna (unidade do Excel) × este fator. É uma
# aproximação (fonte proporcional, não monoespaçada) — propositalmente um pouco
# conservadora, para sobrar espaço em branco em vez de cortar texto.
CHARS_POR_UNIDADE_LARGURA = 1.8


def estimar_linhas(texto, largura_col):
    if not texto:
        return 1
    chars_por_linha = max(10, largura_col * CHARS_POR_UNIDADE_LARGURA)
    linhas = 0
    for paragrafo in str(texto).split("\n"):
        linhas += max(1, math.ceil(len(paragrafo) / chars_por_linha))
    return max(1, linhas)


def calc_row_height(cols, line_pt=14, padding_pt=8, min_pt=20):
    """cols: lista de (texto, largura_da_coluna) para as colunas com wrap_text=True
    que determinam a altura da linha. Devolve a altura (em pontos) que cabe o texto
    mais longo entre elas, nunca menor que min_pt. Usado em vez de altura fixa nas
    abas com texto livre, para não cortar rótulos/observações mais longos."""
    if not cols:
        return min_pt
    linhas = max(estimar_linhas(texto, largura) for texto, largura in cols)
    return max(min_pt, linhas * line_pt + padding_pt)


def build_aba1(wb):
    ws1 = wb.active
    ws1.title = "Cuidados na Alimentação"
    ws1.sheet_view.showGridLines = False

    style_title(ws1, "Cuidados na Alimentação — pós-infarto (25/08/2026)", span=2)
    header_row(ws1, 2, ["Seção", "Orientação"], widths=[26, 95])
    ws1.freeze_panes = "A3"

    rows_aba1 = [
        ("Perfil de risco atual",
         "Homem, 48 anos, IAM em 25/08/2026 (3 procedimentos, FE preservada 72% com "
         "hipocinesia segmentar leve). Dislipidemia mista: LDL 138 (meta pós-IAM <50), "
         "HDL 34 (baixo), triglicerídeos 343 (muito alto), colesterol total 222 (alto). "
         "Placa carotídea leve (1–15%) já mostra aterosclerose além do evento agudo."),
        ("Reduzir",
         "Gordura saturada e gordura trans (frituras, embutidos, carnes gordas, manteiga, "
         "produtos industrializados com 'gordura vegetal hidrogenada')."),
        ("Reduzir",
         "Açúcar e carboidrato refinado — prioridade alta por causa dos triglicerídeos "
         "muito altos (343 mg/dL): refrigerante comum, doces, pão/arroz/massa branca em excesso."),
        ("Reduzir",
         "Sódio — evitar sal em excesso, temperos prontos, embutidos, enlatados e "
         "fast-food; ler rótulo (mg de sódio por porção)."),
        ("Reduzir",
         "Álcool — evitar ou reduzir ao mínimo; álcool piora diretamente os triglicerídeos."),
        ("Priorizar",
         "Fibras: grãos integrais, leguminosas (feijão, lentilha, grão-de-bico), vegetais "
         "e frutas variadas."),
        ("Priorizar",
         "Ômega-3: peixes (preferir 2-3x/semana), azeite de oliva extra virgem, oleaginosas "
         "em pequena quantidade (castanhas, nozes)."),
        ("Priorizar",
         "Proteínas magras: frango sem pele, peixe, ovos, leguminosas; carnes vermelhas com "
         "moderação."),
        ("Atenção especial — triglicerídeos",
         "343 mg/dL é muito alto: o maior impacto individual costuma vir de cortar açúcar "
         "e álcool, não só gordura. Evitar refrigerante comum, suco industrializado, doces "
         "e bebida alcoólica quase por completo até reavaliação médica."),
        ("Atenção especial — LDL",
         "Meta pós-IAM é bem mais rígida que a da população geral, mas varia por diretriz: "
         "<50 mg/dL (diretriz brasileira SBC 2017, usada pelo laboratório), <40 mg/dL "
         "(categoria 'risco extremo' da diretriz SBC 2025, se preencher critérios adicionais) "
         "ou <55 mg/dL (diretrizes ESC/EAS e ACC/AHA para risco muito alto) — pergunte ao "
         "médico qual meta ele está usando. A combinação estatina (Rosucor) + Ezetimiba já "
         "prescrita é sustentada pelo estudo IMPROVE-IT (LDL médio de 53,7 mg/dL com a "
         "combinação, vs. 69,5 mg/dL só com estatina, e menos eventos cardiovasculares). "
         "Fontes na aba 'Fontes e Referências'."),
        ("Aviso pendente — função renal",
         "Há suspeita ainda não confirmada de proteinúria (o exame certo a pedir é a relação "
         "albumina/creatinina urinária, ACR/RAC, padrão KDIGO) e duas elevações de "
         "creatinina/TFG durante a internação (TFG caiu para 57 na véspera da alta) — a "
         "última, aplicando o critério objetivo da KDIGO, já preenche Lesão Renal Aguda "
         "estágio 1 (não é diagnóstico médico, só a fórmula aplicada aos números). O padrão "
         "temporal é compatível com nefropatia por contraste dos 3 cateterismos (ESUR: pico "
         "em 3-5 dias, resolução em 7-14 dias). Enquanto não houver confirmação médica de que "
         "isso se resolveu, o plano NÃO aplica restrição renal (sódio/potássio/proteína) — "
         "apenas o cuidado cardiovascular abaixo. Revisar esta aba assim que o médico "
         "confirmar ou descartar o problema renal."),
        ("Medicações em uso (receita 03/09/2026)",
         "Aspirina Prevent 100mg (1x/dia, almoço) e Ticagrelor 90mg (12/12h, por 1 ano) — "
         "antiagregantes plaquetários; Rosucor (rosuvastatina) 20mg, 2 comprimidos 1x/dia, "
         "e Ezetimiba 10mg 1x/dia — colesterol; Maleato de Enalapril 5mg 12/12h — coração/"
         "pressão (médico alertou para cuidado com pressão baixa). Todas de uso contínuo — "
         "não interromper sem falar com o cardiologista, em especial a dupla Aspirina + "
         "Ticagrelor (risco de trombose no stent)."),
        ("Interação medicamentosa — Ticagrelor",
         "Toranja/grapefruit (fruta ou suco) inibe uma enzima do fígado (CYP3A4) que ajuda a "
         "metabolizar o Ticagrelor. Um estudo publicado mostrou quase o dobro do nível do "
         "remédio no sangue com suco em grande quantidade; a própria bula brasileira do "
         "Brilinta cita esse estudo e diz que não é esperado ser 'clinicamente relevante para "
         "a maioria dos pacientes' — mas recomenda evitar. Como você toma 2 antiagregantes "
         "juntos, a orientação mais segura é evitar toranja/grapefruit. Fontes: aba "
         "'Fontes e Referências', linhas 1-2."),
        ("Interação medicamentosa — Enalapril",
         "Enalapril pode elevar o potássio do sangue (reduz a aldosterona). Evitar sal "
         "light/substitutos de sal (ricos em cloreto de potássio) e suplementos de potássio "
         "sem orientação médica — risco bem documentado de hipercalemia. O potássio (4,9) e a "
         "creatinina (1,50) da alta ficam dentro da faixa considerada tolerável por uma "
         "diretriz de insuficiência cardíaca (ESC 2021) para manter o Enalapril — informação "
         "tranquilizadora, mas essa mesma diretriz recomenda reexame de potássio/creatinina em "
         "1-2 semanas, e reforça evitar sal light por precaução até lá. Fontes: aba 'Fontes e "
         "Referências', linha 3 (Enalapril) e 15 (limiares ESC)."),
        ("Interação medicamentosa — Aspirina + Ticagrelor",
         "A dupla antiagregação (esquema padrão por 1 ano pós-stent, conforme diretriz "
         "ACC/AHA) já aumenta o risco de sangramento; estudos mostram que álcool combinado "
         "com Aspirina aumenta especificamente o sangramento gastrointestinal alto — reforça "
         "a orientação de evitar bebida alcoólica. Fontes: aba 'Fontes e Referências', "
         "linhas 4 e 7."),
        ("Interação medicamentosa — Rosucor (estatina)",
         "Evitar suplementos de 'arroz de levedura vermelha' (red yeast rice): o NIH americano "
         "alerta que contêm uma substância quimicamente idêntica a uma estatina e podem somar "
         "efeitos colaterais/toxicidade com a Rosucor. Já a toranja/grapefruit NÃO tem "
         "interação relevante conhecida com a Rosucor especificamente (ela usa outra via "
         "metabólica, CYP2C9) — o motivo de evitar toranja neste plano é o Ticagrelor, não a "
         "Rosucor. Fontes: aba 'Fontes e Referências', linhas 5-6."),
        ("Reabilitação cardíaca",
         "A diretriz mais recente de síndrome coronariana aguda (ACC/AHA/ACEP/NAEMSP/SCAI "
         "2025) recomenda encaminhamento para reabilitação cardíaca com o grau mais forte de "
         "recomendação (Classe 1), idealmente ainda na alta — reduz mortalidade, reinfarto e "
         "reinternações. Não é sobre alimentação diretamente, mas costuma incluir "
         "acompanhamento nutricional; vale perguntar ao médico se já foi encaminhado. Fonte: "
         "aba 'Fontes e Referências', linha 17."),
        ("Leitura de rótulos",
         "Ao comprar um produto industrializado, olhar sempre: sódio por porção, açúcares "
         "totais/adicionados, e a lista de ingredientes procurando 'gordura hidrogenada' ou "
         "'gordura trans'. Prefira produtos com menos ingredientes e sem açúcar adicionado."),
    ]

    r = 3
    for secao, texto in rows_aba1:
        is_warning = secao.startswith(("Aviso pendente", "Interação medicamentosa", "Medicações em uso"))
        c1 = ws1.cell(row=r, column=1, value=secao)
        c2 = ws1.cell(row=r, column=2, value=texto)
        fill = PatternFill("solid", fgColor=COLOR_WARN_BG if is_warning else COLOR_SECTION_BG)
        c1.font = Font(name=FONT_NAME, size=10.5, bold=True)
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c1.fill = fill
        c1.border = BORDER_ALL
        c2.font = Font(name=FONT_NAME, size=10.5)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.border = BORDER_ALL
        ws1.row_dimensions[r].height = calc_row_height([(secao, 26), (texto, 95)], min_pt=30)
        r += 1


def build_aba2(wb):
    ws2 = wb.create_sheet("Plano Geral - Restaurante")
    ws2.sheet_view.showGridLines = False

    style_title(ws2, "Plano Geral — o que comer, inclusive fora de casa", span=2)
    header_row(ws2, 2, ["Seção", "Conteúdo"], widths=[26, 95])
    ws2.freeze_panes = "A3"

    rows_aba2 = [
        ("Modelo de prato",
         "1/2 do prato: vegetais e salada (à vontade, tempero simples com azeite/limão). "
         "1/4 do prato: proteína magra grelhada/assada (peixe, frango sem pele, ovo, "
         "leguminosas). 1/4 do prato: carboidrato integral (arroz integral, batata-doce, "
         "mandioca, pão de forma integral) em porção moderada."),
        ("O que pedir",
         "Grelhado, assado ou cozido — sem manteiga extra ou molho cremoso. Peça o molho "
         "à parte para controlar a quantidade. Troque frituras por grelhados/assados. "
         "Troque refrigerante comum por água ou refrigerante zero. Sobremesa só ocasional "
         "e em porção pequena (dividir com alguém, se possível)."),
        ("O que evitar",
         "Frituras (empanados, milanesas), molhos à base de creme de leite/manteiga, "
         "embutidos (bacon, linguiça), pratos 'à parmegiana' ou gratinados com muito "
         "queijo, refrigerante comum, sobremesas grandes ou muito açucaradas."),
        ("Perguntas úteis ao garçom",
         "\"Esse prato pode vir grelhado, sem manteiga extra?\" · \"O molho leva creme de "
         "leite ou manteiga?\" · \"Dá para trocar a guarnição (fritas/farofa) por salada ou "
         "legumes?\" · \"Tem opção de refrigerante zero ou água com gás?\""),
        ("Exemplos de pratos que funcionam",
         "Peixe grelhado + legumes salteados + arroz integral. Frango grelhado sem pele + "
         "salada variada. Prato de grãos/leguminosas (feijoada de grão-de-bico, saladas com "
         "lentilha) + vegetais. Omelete de claras/ovo com legumes + salada."),
        ("Em festas/eventos",
         "Priorize proteína grelhada e salada primeiro, para 'preencher' o prato com "
         "opções seguras. Álcool: evitar ou limitar ao mínimo por causa dos "
         "triglicerídeos. Doces: só uma porção pequena, ocasionalmente."),
    ]

    r = 3
    for secao, texto in rows_aba2:
        c1 = ws2.cell(row=r, column=1, value=secao)
        c2 = ws2.cell(row=r, column=2, value=texto)
        c1.font = Font(name=FONT_NAME, size=10.5, bold=True)
        c1.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c1.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
        c1.border = BORDER_ALL
        c2.font = Font(name=FONT_NAME, size=10.5)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.border = BORDER_ALL
        ws2.row_dimensions[r].height = calc_row_height([(secao, 26), (texto, 95)], min_pt=30)
        r += 1


def category_order(alimentos):
    """Ordem das categorias: CATEGORY_ORDER primeiro, categorias novas (que
    não estejam na lista) entram no final, na ordem em que aparecerem."""
    seen_order = list(CATEGORY_ORDER)
    for item in alimentos:
        if item["categoria"] not in seen_order:
            seen_order.append(item["categoria"])
    return seen_order


def load_alimentos():
    with open(JSON_PATH, encoding="utf-8") as f:
        alimentos = json.load(f)

    # ordena por categoria e mantém a ordem original dentro de cada categoria.
    order = category_order(alimentos)
    alimentos_ordenados = sorted(alimentos, key=lambda item: order.index(item["categoria"]))
    return alimentos_ordenados


def _ler_gostoso_de(path):
    """Lê um .xlsx (local ou baixado) e devolve {(alimento, marca): valor de
    "Gostoso"}. Nunca levanta erro — arquivo ausente, aba ausente ou coluna
    "Gostoso" ainda não existente simplesmente devolvem {}."""
    if not path or not os.path.exists(path):
        return {}
    try:
        wb_antigo = openpyxl.load_workbook(path, data_only=True)
        if "Alimentos e Quantidades" not in wb_antigo.sheetnames:
            return {}
        ws_antigo = wb_antigo["Alimentos e Quantidades"]

        col_idx = {}
        for col in range(1, ws_antigo.max_column + 1):
            header = ws_antigo.cell(row=3, column=col).value
            if header in ("Alimento", "Marca", "Gostoso"):
                col_idx[header] = col
        if "Alimento" not in col_idx or "Gostoso" not in col_idx:
            return {}

        previos = {}
        for row in range(4, ws_antigo.max_row + 1):
            alimento = ws_antigo.cell(row=row, column=col_idx["Alimento"]).value
            if not alimento:
                continue  # linha de cabeçalho de categoria ou em branco
            marca = ws_antigo.cell(row=row, column=col_idx["Marca"]).value if "Marca" in col_idx else ""
            gostoso = ws_antigo.cell(row=row, column=col_idx["Gostoso"]).value
            if gostoso:
                previos[(alimento, marca or "")] = gostoso
        return previos
    except Exception:
        return {}


def load_gostoso_previo(sheets_export_path=None):
    """Devolve {(alimento, marca): valor de "Gostoso"} para não perder as
    avaliações pessoais que o usuário já preencheu à mão, quando a planilha for
    regenerada por causa de um alimento novo.

    Lê duas fontes possíveis e funde as duas (a mesma chave (alimento, marca)
    tem os dois valores comparados; se só uma tiver preenchido, usa essa; se
    as duas tiverem preenchido e forem diferentes, `sheets_export_path` vence,
    por ser a cópia mais recente do que o usuário está editando de fato):

    - `sheets_export_path`: uma cópia do Google Sheets baixada como .xlsx
      pouco antes de regenerar (ver planilha/SHEETS_SYNC.md) — é o que
      importa quando o usuário está editando "Gostoso" lá.
    - `OUT_PATH`: o .xlsx local (git) — cobre o caso de quem só usa o arquivo
      enviado pelo chat, sem Google Sheets."""
    previos = _ler_gostoso_de(OUT_PATH)
    previos.update(_ler_gostoso_de(sheets_export_path))
    return previos


def build_aba3(wb, gostoso_previo=None):
    """Colunas: Categoria | Alimento | Marca | Porção | Frequência | Classificação
    | Gostoso | Observação. "Marca" vem do JSON (como porção/classificação — a
    skill preenche). "Gostoso" é avaliação pessoal do usuário, preenchida à mão
    via menu suspenso na planilha — nunca decidida pela skill — e é preservada
    entre regenerações via `gostoso_previo` (ver load_gostoso_previo)."""
    gostoso_previo = gostoso_previo or {}
    ws3 = wb.create_sheet("Alimentos e Quantidades")
    ws3.sheet_view.showGridLines = False

    style_title(ws3, "Alimentos e Quantidades — lista de referência", span=8)

    ws3.row_dimensions[2].height = 20
    legend = [
        (2, "Liberado", COLOR_LIBERADO),
        (3, "Moderar", COLOR_MODERAR),
        (4, "Evitar", COLOR_EVITAR),
    ]
    for col, text, color in legend:
        cell = ws3.cell(row=2, column=col, value=text)
        cell.font = Font(name=FONT_NAME, size=9, bold=True)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_ALL
    gostoso_legend = ws3.cell(row=2, column=7,
                               value="Gostoso: preencha você mesmo (menu suspenso na coluna)")
    gostoso_legend.font = Font(name=FONT_NAME, size=9, italic=True)
    gostoso_legend.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)

    headers3 = ["Categoria", "Alimento", "Marca", "Porção", "Frequência",
                "Classificação", "Gostoso", "Observação"]
    widths3 = [16, 24, 14, 16, 18, 15, 13, 40]
    header_row(ws3, 3, headers3, widths=widths3)
    ws3.freeze_panes = "A4"

    alimentos = load_alimentos()

    dv_gostoso = DataValidation(
        type="list",
        formula1=f'"{",".join(GOSTOSO_OPCOES)}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
    )
    dv_gostoso.error = "Escolha uma opção do menu (ou deixe em branco)."
    dv_gostoso.errorTitle = "Fora da lista"
    dv_gostoso.errorStyle = "warning"
    ws3.add_data_validation(dv_gostoso)

    r = 4
    current_cat = None
    for item in alimentos:
        categoria = item["categoria"]
        if categoria != current_cat:
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            cat_cell = ws3.cell(row=r, column=1, value=categoria)
            cat_cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=COLOR_CATEGORY_FG)
            cat_cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY_BG)
            cat_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col in range(1, 9):
                ws3.cell(row=r, column=col).border = BORDER_ALL
            ws3.row_dimensions[r].height = 20
            current_cat = categoria
            r += 1

        classif = item["classificacao"]
        marca = item.get("marca", "")
        gostoso = gostoso_previo.get((item["alimento"], marca), "")
        values = [None, item["alimento"], marca, item["porcao"], item["frequencia"],
                  classif, gostoso, item["observacao"]]
        for col, val in enumerate(values, start=1):
            if col == 1:
                continue
            cell = ws3.cell(row=r, column=col, value=val)
            cell.font = Font(name=FONT_NAME, size=10.5, bold=(col == 6))
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4, 5, 6, 7) else "left",
                vertical="center", wrap_text=True, indent=(1 if col in (2, 8) else 0)
            )
            cell.border = BORDER_ALL
            if col == 6:
                cell.fill = PatternFill("solid", fgColor=CLASS_COLORS.get(classif, "FFFFFF"))
            if col == 7 and val:
                cell.fill = PatternFill("solid", fgColor=GOSTOSO_COLORS.get(val, "FFFFFF"))
        dv_gostoso.add(f"G{r}")
        ws3.row_dimensions[r].height = calc_row_height(
            [(item["alimento"], 24), (item["observacao"], 40)], min_pt=30)
        r += 1

    r += 1
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    note_texto = ("Lista com os alimentos acrescentados até agora — não é uma lista "
                  "completa. Use a skill 'adicionar-alimento' (ex.: \"adicione "
                  "sucrilhos\") para ir acrescentando novos itens aos poucos. A coluna "
                  "'Gostoso' é sua — escolha do menu suspenso; o valor é preservado "
                  "mesmo quando a planilha é regenerada por um alimento novo.")
    note = ws3.cell(row=r, column=1, value=note_texto)
    note.font = Font(name=FONT_NAME, size=9.5, italic=True)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.row_dimensions[r].height = calc_row_height([(note_texto, sum(widths3))], min_pt=30)


# Cada linha: (afirmação que essa fonte sustenta, nome da fonte, URL, tipo de fonte, área).
# Área = "Alimentação" (interação medicamento x alimento ou diretriz de dieta) ou "Clínico
# geral" (contexto clínico mais amplo, sem relação direta com o que comer). Pesquisado em
# 04/09/2026 (itens 1-10) e reaproveitado de uma reavaliação clínica mais ampla feita numa
# sessão separada (itens 11-22). Nenhuma destas fontes substitui a avaliação do médico — são
# a base do que está escrito na aba "Cuidados na Alimentação" e em
# referencia/historico-tratamento.md, para conferência do paciente.
#
# IMPORTANTE: os textos das abas 1/4 citam estas fontes por número de linha (ex.: "linha 15").
# Ao editar, prefira ACRESCENTAR no final em vez de reordenar — reordenar quebra essas
# referências cruzadas (procure por "Fontes e Referências', linha" nas outras funções deste
# arquivo antes de mudar a posição de um item existente).
FONTES = [
    ("Suco de toranja em quantidade alta quase dobra o nível de Ticagrelor no sangue e "
     "aumenta seu efeito antiagregante",
     "Holmberg et al., 2013, Br J Clin Pharmacol (PubMed)",
     "https://pubmed.ncbi.nlm.nih.gov/23126367/",
     "Estudo clínico revisado por pares", "Alimentação"),
    ("Bula brasileira do Brilinta (Ticagrelor) cita esse estudo e orienta evitar suco de "
     "toranja em grande quantidade",
     "Bula profissional Brilinta, AstraZeneca do Brasil (aprovada pela Anvisa)",
     "https://www.azmed.com.br/content/dam/multibrand/br/pt/azmed-2022/home/bulas-profissionais/bulas/Brilinta_Bula_Profissional.pdf",
     "Bula oficial (Anvisa)", "Alimentação"),
    ("Enalapril (IECA) + sal light/suplemento de potássio aumenta risco de hipercalemia",
     "StatPearls (NCBI Bookshelf), \"Enalapril\"",
     "https://www.ncbi.nlm.nih.gov/books/NBK557708/",
     "Referência clínica revisada por pares", "Alimentação"),
    ("Álcool combinado com Aspirina aumenta o risco de sangramento gastrointestinal alto",
     "Kaufman et al., 1999, Ann Epidemiol (PubMed)",
     "https://pubmed.ncbi.nlm.nih.gov/10566713/",
     "Estudo epidemiológico revisado por pares", "Alimentação"),
    ("Arroz de levedura vermelha (red yeast rice) pode causar os mesmos efeitos colaterais e "
     "interações de uma estatina",
     "NIH — National Center for Complementary and Integrative Health",
     "https://www.nccih.nih.gov/health/red-yeast-rice",
     "Órgão oficial do governo dos EUA (NIH)", "Alimentação"),
    ("Rosuvastatina não tem interação relevante com toranja (via CYP2C9, diferente da via "
     "CYP3A4 usada pelo Ticagrelor)",
     "Bailey DG et al., 2013, CMAJ, \"Grapefruit-medication interactions\"",
     "https://pmc.ncbi.nlm.nih.gov/articles/PMC3589309/",
     "Revisão científica revisada por pares", "Alimentação"),
    ("Duração de 12 meses de dupla antiagregação (Aspirina + Ticagrelor) após stent em "
     "síndrome coronariana aguda",
     "2016 ACC/AHA Guideline Focused Update on DAPT; reafirmado na diretriz ACS 2025 (JACC)",
     "https://www.acc.org/latest-in-cardiology/ten-points-to-remember/2016/03/25/14/56/2016-acc-aha-guideline-focused-update-on-dapt",
     "Diretriz de sociedades médicas (ACC/AHA)", "Alimentação"),
    ("Metas de LDL pós-infarto: <50 mg/dL (SBC 2017); <40 mg/dL na categoria \"risco extremo\" "
     "(SBC 2025, com critérios adicionais); <55 mg/dL (ESC/EAS 2019 e ACC/AHA, risco muito alto)",
     "Diretriz Brasileira de Dislipidemias e Prevenção da Aterosclerose 2025 (SBC), Arq Bras Cardiol",
     "https://www.scielo.br/j/abc/a/tRJrwGzKX6C4GvMqdJpZcGk/?lang=pt",
     "Diretriz de sociedade médica (SBC)", "Alimentação"),
    ("Diretriz alimentar cardioprotetora: sódio controlado, gordura saturada <6% das "
     "calorias, padrão Mediterrâneo/DASH",
     "American Heart Association — AHA Diet and Lifestyle Recommendations / 2026 Dietary "
     "Guidance to Improve Cardiovascular Health (Circulation)",
     "https://www.heart.org/en/healthy-living/healthy-eating/eat-smart/nutrition-basics/aha-diet-and-lifestyle-recommendations",
     "Diretriz de sociedade médica (AHA)", "Alimentação"),
    ("No Brasil, a Dieta Cardioprotetora Brasileira (DICA Br) aplica essas mesmas "
     "recomendações a alimentos típicos brasileiros",
     "I Diretriz Brasileira de Prevenção Cardiovascular (SBC), Arq Bras Cardiol",
     "https://www.scielo.br/j/abc/a/X94tMKwdnBjkCzVKpXwBqmD/?lang=pt",
     "Diretriz de sociedade médica (SBC)", "Alimentação"),
    ("Critério objetivo de Lesão Renal Aguda (LRA) estágio 1: creatinina sobe ≥0,3 mg/dL em 48h",
     "KDIGO — Clinical Practice Guideline for Acute Kidney Injury, 2012",
     "https://kdigo.org/wp-content/uploads/2016/10/KDIGO-2012-AKI-Guideline-English.pdf",
     "Diretriz de sociedade médica (KDIGO)", "Clínico geral"),
    ("Exame padrão para proteinúria é a relação albumina/creatinina urinária (ACR/RAC), "
     "categorias A1/A2/A3",
     "KDIGO — Clinical Practice Guideline for CKD, 2012 (categorias mantidas na atualização 2024)",
     "https://kdigo.org/wp-content/uploads/2017/02/KDIGO_2012_CKD_GL.pdf",
     "Diretriz de sociedade médica (KDIGO)", "Clínico geral"),
    ("Múltiplas exposições a contraste em dias próximos são fator de risco para lesão renal "
     "por contraste, com pico em 3-5 dias e resolução em 7-14 dias",
     "ESUR — Post-contrast acute kidney injury, Guidelines v10.0, 2018",
     "https://www.esur.org/wp-content/uploads/2022/03/ESUR-Guidelines-10_0-Final-Version.pdf",
     "Diretriz de sociedade médica (ESUR)", "Clínico geral"),
    ("Definição de infarto periprocedimento (tipo 4a): troponina >5x o limite superior + "
     "evidência de nova isquemia",
     "Thygesen K, Alpert JS et al. — Fourth Universal Definition of Myocardial Infarction, 2018",
     "https://www.ahajournals.org/doi/10.1161/CIR.0000000000000617",
     "Documento conjunto de sociedades médicas (ESC/ACC/AHA/WHF)", "Clínico geral"),
    ("Limiares de segurança para manter um IECA (Enalapril): potássio até 5,5 mmol/L e "
     "creatinina com alta até 50% (abaixo de 3 mg/dL) são toleráveis; reexame em 1-2 semanas",
     "ESC — Guidelines for the diagnosis and treatment of acute and chronic heart failure, 2021",
     "https://academic.oup.com/eurheartj/article/42/36/3599/6358045",
     "Diretriz de sociedade médica (ESC)", "Clínico geral"),
    ("Reforço da duração de 12 meses de dupla antiagregação após síndrome coronariana aguda",
     "ESC — Guidelines for the management of acute coronary syndromes, 2023",
     "https://academic.oup.com/eurheartj/article/45/14/1193/7516285",
     "Diretriz de sociedade médica (ESC)", "Clínico geral"),
    ("Encaminhamento para reabilitação cardíaca antes da alta é recomendação Classe 1 (a "
     "mais forte) após síndrome coronariana aguda",
     "ACC/AHA/ACEP/NAEMSP/SCAI — Guideline for the Management of Patients With Acute "
     "Coronary Syndromes, 2025",
     "https://www.ahajournals.org/doi/10.1161/CIR.0000000000001309",
     "Diretriz de sociedades médicas (ACC/AHA e outras)", "Clínico geral"),
    ("Estudo que sustenta a combinação estatina + ezetimiba para reduzir LDL após síndrome "
     "coronariana aguda (IMPROVE-IT)",
     "Cannon CP et al. — Ezetimibe Added to Statin Therapy after Acute Coronary Syndromes, "
     "N Engl J Med 2015",
     "https://www.nejm.org/doi/full/10.1056/NEJMoa1410489",
     "Estudo clínico revisado por pares", "Alimentação"),
    ("HbA1c ≥6,5% sugere diabetes prévio; abaixo de 5,7% (caso do paciente) é consistente com "
     "hiperglicemia de estresse, não diabetes",
     "American Diabetes Association — Standards of Care in Diabetes, edições 2024-2026",
     "https://diabetesjournals.org/care/article/49/Supplement_1/S339/163925/16-Diabetes-Care-in-the-Hospital-Standards-of-Care",
     "Diretriz de sociedade médica (ADA)", "Clínico geral"),
    ("Até 60% dos pacientes com hiperglicemia de estresse na internação desenvolvem diabetes "
     "em 6-12 meses",
     "Sociedade Brasileira de Diabetes — Hiperglicemia Hospitalar no Paciente Não-Crítico",
     "https://diretriz.diabetes.org.br/hiperglicemia-hospitalar-em-paciente-nao-critico/",
     "Diretriz de sociedade médica (SBD)", "Clínico geral"),
    ("Placa carotídea com estenose abaixo de 50% é tratada de forma conservadora "
     "(antiagregante, estatina, controle de pressão), sem indicação de cirurgia",
     "ESC/ESVS — Guidelines on Peripheral Arterial Diseases (2017/2018) e atualização sobre "
     "doença carotídea/vertebral (2023)",
     "https://academic.oup.com/eurheartj/article/39/9/763/4095038",
     "Diretriz de sociedade médica (ESC/ESVS)", "Clínico geral"),
    ("Mesma conduta conservadora para placa carotídea leve, do lado da prevenção de AVC",
     "AHA/ASA — Guideline for the Prevention of Stroke in Patients With Stroke and TIA, 2021",
     "https://www.ahajournals.org/doi/10.1161/STR.0000000000000375",
     "Diretriz de sociedade médica (AHA/ASA)", "Clínico geral"),
]


def build_aba_fontes(wb):
    ws = wb.create_sheet("Fontes e Referências")
    ws.sheet_view.showGridLines = False

    style_title(ws, "Fontes e Referências — de onde vieram as informações médicas", span=6)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    instr = ws.cell(
        row=2, column=1,
        value="Pesquisa feita em 04/09/2026, priorizando bulas oficiais (Anvisa), estudos "
              "revisados por pares (PubMed/PMC), órgãos de saúde do governo (NIH) e diretrizes "
              "de sociedades médicas (SBC, AHA, ACC/ESC, KDIGO, ESUR, ESVS, ADA/SBD). A coluna "
              "'Área' diz se a fonte é sobre alimentação/interação com alimento ou sobre o "
              "quadro clínico em geral. Nenhuma destas fontes substitui a avaliação do seu "
              "médico — use como ponto de partida para perguntar e confirmar na consulta. "
              "Clique no link da coluna 'Fonte' para abrir a página original."
    )
    instr.font = Font(name=FONT_NAME, size=9.5, italic=True)
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 56

    headers = ["Nº", "Afirmação que a fonte sustenta", "Fonte", "Tipo de fonte", "Área", "URL completa"]
    widths = [5, 44, 38, 24, 13, 50]
    header_row(ws, 3, headers, widths=widths)
    ws.freeze_panes = "A4"

    font_normal = Font(name=FONT_NAME, size=10)
    font_link = Font(name=FONT_NAME, size=10, underline="single", color="1155CC")
    font_url = Font(name=FONT_NAME, size=9, color="666666")

    r = 4
    for i, (afirmacao, fonte, url, tipo, area) in enumerate(FONTES, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=afirmacao)
        c3 = ws.cell(row=r, column=3, value=fonte)
        c3.hyperlink = url
        ws.cell(row=r, column=4, value=tipo)
        ws.cell(row=r, column=5, value=area)
        ws.cell(row=r, column=6, value=url)
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = {3: font_link, 6: font_url}.get(col, font_normal)
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 5) else "left",
                vertical="center", wrap_text=True, indent=(1 if col in (2, 3, 4, 6) else 0)
            )
            cell.border = BORDER_ALL
        ws.row_dimensions[r].height = calc_row_height(
            [(afirmacao, 44), (fonte, 38), (url, 50)], min_pt=30)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    note = ws.cell(
        row=r, column=1,
        value="Lista completa com os mesmos links (mais fácil de clicar em Markdown) também em "
              "referencia/historico-tratamento.md, seção 'Referências médicas'."
    )
    note.font = Font(name=FONT_NAME, size=9.5, italic=True)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 24


def build_ref_alimentos(wb):
    """Aba auxiliar (oculta) que alimenta as fórmulas e os menus suspensos da
    aba 'Montar Refeição': a tabela completa de alimentos, listas por
    refeição já filtradas/ordenadas (Liberado > Moderar > Evitar), uma
    tabela de "alimento em destaque" por categoria+refeição e a tabela de
    categorias que combinam bem entre si. Tudo com valores fixos (sem
    fórmula) — é recalculada do zero a cada vez que este script roda."""
    ws = wb.create_sheet("Ref_Alimentos")
    ws.sheet_state = "hidden"

    alimentos = load_alimentos()
    categorias = category_order(alimentos)

    headers = ["Alimento", "Categoria", "Classificação", "Porção", "Frequência",
               "Observação"] + MEALS
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    for idx, item in enumerate(alimentos, start=2):
        ws.cell(row=idx, column=1, value=item["alimento"])
        ws.cell(row=idx, column=2, value=item["categoria"])
        ws.cell(row=idx, column=3, value=item["classificacao"])
        ws.cell(row=idx, column=4, value=item["porcao"])
        ws.cell(row=idx, column=5, value=item["frequencia"])
        ws.cell(row=idx, column=6, value=item["observacao"])
        refeicoes = set(item.get("refeicoes", []))
        for j, meal in enumerate(MEALS, start=7):
            ws.cell(row=idx, column=j, value="X" if meal in refeicoes else "")

    # Listas por refeição (colunas L, M, N, O = 12-15), para os menus
    # suspensos: só os alimentos elegíveis para aquela refeição, com
    # Liberado primeiro, depois Moderar, depois Evitar.
    tier = {"Liberado": 0, "Moderar": 1, "Evitar": 2}
    list_col = {"Café da manhã": 12, "Almoço": 13, "Lanche": 14, "Jantar": 15}
    for meal, col in list_col.items():
        ws.cell(row=1, column=col, value=f"Lista — {meal}")
        elegiveis = [it for it in alimentos if meal in set(it.get("refeicoes", []))]
        elegiveis.sort(key=lambda it: tier.get(it["classificacao"], 9))
        for i, it in enumerate(elegiveis, start=2):
            ws.cell(row=i, column=col, value=it["alimento"])

        # Nome definido apontando para a lista — é isso que o menu suspenso
        # da aba "Montar Refeição" usa (ver MEAL_DEFINED_NAME).
        letter = get_column_letter(col)
        name = MEAL_DEFINED_NAME[meal]
        wb.defined_names[name] = DefinedName(name, attr_text=f"Ref_Alimentos!${letter}$2:${letter}$101")

    # Tabela "alimento em destaque" por categoria+refeição (colunas Q, R =
    # 17, 18): usada pela coluna "Sugestão". Preferência por um item
    # Liberado; se não houver nenhum, usa o primeiro elegível da categoria.
    ws.cell(row=1, column=17, value="Chave (Categoria|Refeição)")
    ws.cell(row=1, column=18, value="Alimento em destaque")
    row = 2
    for meal in MEALS:
        for categoria in categorias:
            candidatos = [it for it in alimentos
                          if it["categoria"] == categoria and meal in set(it.get("refeicoes", []))]
            if not candidatos:
                continue
            liberados = [it for it in candidatos if it["classificacao"] == "Liberado"]
            destaque = liberados[0] if liberados else candidatos[0]
            ws.cell(row=row, column=17, value=f"{categoria}|{meal}")
            ws.cell(row=row, column=18, value=destaque["alimento"])
            row += 1

    # Tabela de categorias que combinam bem (colunas T, U = 20, 21).
    ws.cell(row=1, column=20, value="Categoria escolhida")
    ws.cell(row=1, column=21, value="Categoria sugerida")
    for i, (origem, sugerida) in enumerate(CATEGORIA_COMBINA.items(), start=2):
        ws.cell(row=i, column=20, value=origem)
        ws.cell(row=i, column=21, value=sugerida)

    return ws


def _build_meal_block(ws, start_row, meal_name, n_items=5):
    """Escreve um bloco de refeição (título + cabeçalho + N linhas de item)
    a partir de start_row e devolve a próxima linha livre depois do bloco.

    Altura de linha fica fixa aqui (ao contrário das outras abas) porque
    Categoria/Sugestão/Conflito são preenchidas por FÓRMULA — o texto real só
    existe depois que o Excel calcula, então não dá para estimar o tamanho na
    hora de gerar o arquivo com calc_row_height()."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    title_cell = ws.cell(row=start_row, column=1, value=meal_name.upper())
    title_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_CATEGORY_FG)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, 9):
        ws.cell(row=start_row, column=col).border = BORDER_ALL
    ws.row_dimensions[start_row].height = 20

    header_r = start_row + 1
    headers = ["Nº", "Alimento", "Categoria", "Porção sugerida", "Quantidade a comer",
               "Classificação", "Sugestão (com base no item anterior)", "Conflito"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_r, column=i, value=h)
        cell.font = Font(name=FONT_NAME, size=9.5, bold=True, color=COLOR_TITLE_FG)
        cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[header_r].height = 30

    first_row = header_r + 1
    last_row = first_row + n_items - 1

    for offset in range(n_items):
        r = first_row + offset

        n_cell = ws.cell(row=r, column=1, value=offset + 1)
        n_cell.font = Font(name=FONT_NAME, size=10.5)
        n_cell.alignment = Alignment(horizontal="center", vertical="center")

        b_cell = ws.cell(row=r, column=2)
        b_cell.font = Font(name=FONT_NAME, size=10.5)
        b_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_cell = ws.cell(
            row=r, column=3,
            value=f'=IFERROR(INDEX(Ref_Alimentos!$B$2:$B$101,MATCH($B{r},Ref_Alimentos!$A$2:$A$101,0)),"")'
        )
        c_cell.font = Font(name=FONT_NAME, size=10, italic=True, color="666666")
        c_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        d_cell = ws.cell(
            row=r, column=4,
            value=f'=IFERROR(INDEX(Ref_Alimentos!$D$2:$D$101,MATCH($B{r},Ref_Alimentos!$A$2:$A$101,0)),"")'
        )
        d_cell.font = Font(name=FONT_NAME, size=10, italic=True, color="666666")
        d_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        e_cell = ws.cell(row=r, column=5)
        e_cell.font = Font(name=FONT_NAME, size=10.5)
        e_cell.alignment = Alignment(horizontal="center", vertical="center")

        f_cell = ws.cell(
            row=r, column=6,
            value=f'=IFERROR(INDEX(Ref_Alimentos!$C$2:$C$101,MATCH($B{r},Ref_Alimentos!$A$2:$A$101,0)),"")'
        )
        f_cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
        f_cell.alignment = Alignment(horizontal="center", vertical="center")

        if offset == 0:
            g_cell = ws.cell(row=r, column=7, value="")
        else:
            prev_cat = f"$C{r - 1}"
            formula = (
                f'=IF($B{r - 1}="","",'
                f'IFERROR(INDEX(Ref_Alimentos!$R$2:$R$60,MATCH('
                f'IFERROR(INDEX(Ref_Alimentos!$U$2:$U$9,MATCH({prev_cat},Ref_Alimentos!$T$2:$T$9,0)),"")'
                f'&"|"&"{meal_name}",Ref_Alimentos!$Q$2:$Q$60,0)),""))'
            )
            g_cell = ws.cell(row=r, column=7, value=formula)
        g_cell.font = Font(name=FONT_NAME, size=10, italic=True, color=COLOR_TITLE_BG)
        g_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        b_rng = f"$B${first_row}:$B${last_row}"
        f_rng = f"$F${first_row}:$F${last_row}"
        conflito_formula = (
            f'=IF($B{r}="","",IF($F{r}="Evitar","⚠ Conflito",'
            f'IF(AND($F{r}="Moderar",SUMPRODUCT(({b_rng}<>"")*(ROW({b_rng})<>ROW($B{r}))*'
            f'(({f_rng}="Moderar")+({f_rng}="Evitar")))>0),"⚠ Conflito","")))'
        )
        h_cell = ws.cell(row=r, column=8, value=conflito_formula)
        h_cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
        h_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER_ALL
        ws.row_dimensions[r].height = 26

    # Uma lista de validação não pode apontar direto para um intervalo em
    # outra aba — só para um nome definido (ou um intervalo na mesma aba).
    # Por isso o formula1 usa o nome definido, não "Ref_Alimentos!...".
    dv = DataValidation(
        type="list",
        formula1=f"={MEAL_DEFINED_NAME[meal_name]}",
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        showErrorMessage=True,
    )
    dv.error = "Escolha um alimento da lista da coluna (ou deixe em branco)."
    dv.errorTitle = "Alimento fora da lista"
    dv.errorStyle = "warning"
    dv.promptTitle = f"Alimentos para {meal_name}"
    dv.prompt = "Lista filtrada para esta refeição, com os alimentos Liberados primeiro."
    ws.add_data_validation(dv)
    dv.add(f"B{first_row}:B{last_row}")

    class_range = f"F{first_row}:F{last_row}"
    ws.conditional_formatting.add(
        class_range, CellIsRule(operator="equal", formula=['"Liberado"'],
                                 fill=PatternFill("solid", fgColor=COLOR_LIBERADO)))
    ws.conditional_formatting.add(
        class_range, CellIsRule(operator="equal", formula=['"Moderar"'],
                                 fill=PatternFill("solid", fgColor=COLOR_MODERAR)))
    ws.conditional_formatting.add(
        class_range, CellIsRule(operator="equal", formula=['"Evitar"'],
                                 fill=PatternFill("solid", fgColor=COLOR_EVITAR)))

    conflito_range = f"H{first_row}:H{last_row}"
    ws.conditional_formatting.add(
        conflito_range, CellIsRule(operator="equal", formula=['"⚠ Conflito"'],
                                    fill=PatternFill("solid", fgColor=COLOR_WARN_BG),
                                    font=Font(name=FONT_NAME, size=10.5, bold=True, color="C0392B")))

    return last_row + 2  # próxima linha livre, com 1 linha de espaço


def build_aba4(wb):
    ws4 = wb.create_sheet("Montar Refeição")
    ws4.sheet_view.showGridLines = False

    widths = [5, 30, 16, 16, 20, 14, 34, 14]
    for i, w in enumerate(widths, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    style_title(ws4, "Montar uma Refeição — escolha os alimentos de cada refeição", span=8, row=1)

    ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    instr = ws4.cell(
        row=2, column=1,
        value="Escolha o alimento no menu suspenso da coluna 'Alimento' — a lista já vem filtrada "
              "por refeição (por isso não aparece arroz/feijão no café da manhã, por exemplo) e com "
              "os alimentos Liberados primeiro. Categoria, Porção sugerida e Classificação preenchem "
              "sozinhas. 'Sugestão' mostra um alimento que costuma combinar bem com o item escolhido "
              "logo acima (é uma dica, não é obrigatório escolher). Se 'Conflito' mostrar "
              "'⚠ Conflito', evite reunir esses itens na mesma refeição."
    )
    instr.font = Font(name=FONT_NAME, size=9.5, italic=True)
    instr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[2].height = 48

    ws4.row_dimensions[3].height = 20
    legend = [
        (1, "Legenda:", None),
        (2, "Liberado", COLOR_LIBERADO),
        (3, "Moderar", COLOR_MODERAR),
        (4, "Evitar", COLOR_EVITAR),
        (6, "⚠ Conflito", COLOR_WARN_BG),
    ]
    for col, text, color in legend:
        cell = ws4.cell(row=3, column=col, value=text)
        cell.font = Font(name=FONT_NAME, size=9, bold=True,
                          color="C0392B" if text == "⚠ Conflito" else "000000")
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 5
    for meal in MEALS:
        row = _build_meal_block(ws4, row, meal)

    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    note = ws4.cell(
        row=row, column=1,
        value="Limitações conhecidas: a ordem do menu suspenso é fixa (Liberado > Moderar > Evitar), "
              "ela não se reordena sozinha conforme o item anterior — quem indica o que combina "
              "melhor é a coluna 'Sugestão'. A coluna 'Conflito' é uma aproximação (marca um item "
              "'Evitar' sozinho, ou dois itens 'Moderar'/'Evitar' juntos na mesma refeição) — não é "
              "uma checagem real de interação entre os dois alimentos específicos."
    )
    note.font = Font(name=FONT_NAME, size=9, italic=True)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[row].height = 44


def main():
    # Se GOSTOSO_SHEETS_EXPORT apontar para um .xlsx baixado do Google Sheets
    # (ver planilha/SHEETS_SYNC.md), os valores de "Gostoso" preenchidos lá
    # têm prioridade sobre os do .xlsx local ao fundir.
    gostoso_previo = load_gostoso_previo(os.environ.get("GOSTOSO_SHEETS_EXPORT"))

    wb = openpyxl.Workbook()
    build_aba1(wb)
    build_aba2(wb)
    build_aba3(wb, gostoso_previo)
    build_aba4(wb)
    build_aba_fontes(wb)
    build_ref_alimentos(wb)
    wb.save(OUT_PATH)
    print("Salvo em", OUT_PATH)


if __name__ == "__main__":
    main()
